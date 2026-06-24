from rest_framework.generics import ListAPIView
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from django.db.models import QuerySet, Sum
from django.db import connection
from sheets import customPermissions
from django.http import HttpResponse
from collections import OrderedDict
from django.db.models import Q
from datetime import date
from itertools import groupby
from django.db.models import Count, Q, Sum
from django.utils import timezone


import jdatetime as jdt
import pandas as pd
import io
import re
from io import BytesIO

from sheets.models import Project, Sheet, User, Food_data, Report, DailyReportSetting, current_mont_days, get_persian_weekday
from esfa_eyes.models import EsfaEyes
from sheets.serializers import ProjectSerializer, SheetSerializer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

def camel_to_snake(s: str) -> str:
    pattern = re.compile(r"(?<!^)(?=[A-Z])")
    snake = pattern.sub("_", s).lower()
    return snake


class ProjectListApiView(ListAPIView):

    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [permissions.IsAuthenticated]


class SheetApiView(APIView):

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, year: str, month: str):
        sheet, created = Sheet.objects.get_or_create(
            user=self.request.user, year=year, month=month
        )
        if created:
            sheet.setup_sheet()

        # Self-heal sheets that somehow lost their rows (legacy/corrupt data, an old
        # bug, etc.). Serving an empty grid would make the page unusable and could
        # lead to an empty save, so rebuild a blank month for this period.
        if not sheet.data:
            sheet.normalize_sheet()

        serializer = SheetSerializer(sheet)
        data = dict(serializer.data)
        # Manager comments are internal verifier notes and should not be exposed on
        # the employee-facing hours page.
        data.pop("manager_level_1_comment", None)
        data.pop("manager_level_2_comment", None)
        return Response(data, status=status.HTTP_200_OK)

    def post(self, request, year: str, month: str):
        if "saveSheet" in request.data:
            user = self.request.user
            sheet, created = Sheet.objects.get_or_create(
                user=user, year=year, month=month
            )
            if created:
                sheet.setup_sheet()

            if sheet.submitted:
                return Response(
                    {"error": "Submitted sheets cannot be edited unless they are rejected first."},
                    status=status.HTTP_409_CONFLICT,
                )

            data = request.data.get("data", [])
            # Guard against accidentally emptying a sheet. The hours grid always has
            # exactly one row per day and there is no UI to add or remove rows, so a
            # payload that is empty or has fewer rows than what is already stored is
            # never a legitimate edit -- it almost always means a dropped connection,
            # an incomplete page load, or a client-side bug. Reject it instead of
            # overwriting good data.
            existing_rows = len(sheet.data or [])
            if not isinstance(data, list) or len(data) == 0 or len(data) < existing_rows:
                return Response(
                    {
                        "error": (
                            "Sheet data looks incomplete, so it was not saved to avoid "
                            "overwriting your hours. Please reload the page and try again."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            data.sort(key=lambda row: int(row.get("Day", 0)))
            sheet.data = data
            sheet.normalize_sheet()
            return Response({"success": True}, status=status.HTTP_200_OK)
        elif "editSheet" in request.data:
            # `editSheet` uses QuerySet.update(), which writes raw SQL and bypasses
            # Sheet.save() and every safeguard in it. Restrict it to a fixed set of
            # scalar payment/note fields so it can never blank the hours grid (e.g.
            # via field="data") or tamper with submission/verification state.
            EDITABLE_FIELDS = {
                "wage", "base_payment", "reduction1", "reduction2", "reduction3",
                "addition1", "addition2", "food_reduction", "payment_status",
                "note_hours",
            }
            row = request.data.get("row", {})
            field = request.data.get("field", "")
            db_field = camel_to_snake(field)
            if db_field not in EDITABLE_FIELDS:
                return Response(
                    {"error": f"Field '{field}' cannot be edited from this endpoint."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            user = User.objects.get(pk=row["userID"])
            sheets = Sheet.objects.filter(user=user, year=year, month=month)
            sheets.update(**{db_field: row[field]})
            first = sheets.first()
            if first is None:
                return Response(
                    {"error": "Sheet not found."}, status=status.HTTP_404_NOT_FOUND
                )
            return Response(first.get_payment_info(), status=status.HTTP_200_OK)

    def put(self, request, year: str, month: str):
        try:
            sheet = Sheet.objects.get(user=self.request.user, year=year, month=month)
        except Sheet.DoesNotExist:
            return Response({"notFound": True}, status=status.HTTP_404_NOT_FOUND)
        if sheet.submitted:
            return Response(
                {"error": "Sheet is already submitted and cannot be unsubmitted by the user."},
                status=status.HTTP_409_CONFLICT,
            )

        if (
            request.user.check_info()
        ):  # if the user has entered needed personal information
            sheet.submitted = True
            # A fresh submission/re-submission starts the approval chain again.
            sheet.manager_level_1_verified = False
            sheet.manager_level_2_verified = False
            sheet.supreme_verified = False
            sheet.manager_level_1_verified_at = None
            sheet.manager_level_2_verified_at = None
            sheet.supreme_verified_at = None
            sheet.manager_level_1_verified_by = None
            sheet.manager_level_2_verified_by = None
            sheet.supreme_verified_by = None
            sheet.last_rejected_by = None
            sheet.last_rejected_at = None
            sheet.rejection_reason = ""
            sheet.sync_legacy_verification_fields()
            sheet.save()
            return Response({"success": True}, status=status.HTTP_200_OK)
        return Response({"flaw": True}, status=status.HTTP_200_OK)


class InfoApiView(APIView):

    permission_classes = [permissions.IsAuthenticated]

    INFO_SHEET_FIELDS = ["id", "user", "year", "month", "data", "total"]
    NON_PROJECT_COLUMNS = {
        "Day", "WeekDay", "Hours", "Auto Hours", "Remote", "Rest", "Total",
        "Attendance", "Description", "Note Hours",
    }

    def get(self, request):

        today = jdt.date.today()
        month, year = today.month, today.year

        # Keep this page fast by selecting only fields used by the charts.  Do not
        # select_related('user') here because this page must keep working even when
        # optional approval User fields exist in code but their DB migration has not
        # been applied yet.
        base_sheets = Sheet.objects.only(*self.INFO_SHEET_FIELDS).filter(user__is_active=True)
        user_sheets = base_sheets.filter(user=request.user)

        user_month_info = self.get_info(
            user_sheets.filter(year=year, month=month)
        )
        user_year_info = self.get_info(
            user_sheets.filter(year=year).exclude(month=month)
        )
        user_year_info = user_year_info.add(user_month_info, fill_value=0)
        user_tot_info = self.get_info(
            user_sheets.exclude(year=year)
        )
        user_tot_info = user_tot_info.add(user_year_info, fill_value=0)

        esfa_month_info = self.get_info(base_sheets.filter(year=year, month=month))
        esfa_year_info = self.get_info(
            base_sheets.filter(year=year).exclude(month=month)
        )
        esfa_year_info = esfa_year_info.add(esfa_month_info, fill_value=0)
        esfa_tot_info = self.get_info(base_sheets.exclude(year=year))
        esfa_tot_info = esfa_tot_info.add(esfa_year_info, fill_value=0)

        last_month = month - 1 if month != 1 else 12
        last_month_year = year if month != 1 else year - 1
        monthly_sheets = (
            user_sheets.filter(year=year)
            .values("total", "month")
            .order_by("month")
        )
        user_monthly_hours = list(monthly_sheets)
        info = {
            "user_month_info": user_month_info.to_dict(),
            "user_year_info": user_year_info.to_dict(),
            "user_tot_info": user_tot_info.to_dict(),
            "esfa_month_info": esfa_month_info.to_dict(),
            "esfa_year_info": esfa_year_info.to_dict(),
            "esfa_tot_info": esfa_tot_info.to_dict(),
            "last_hero": self.get_hero(last_month_year, last_month),
            "last_esfa_mean": self.get_month_mean(last_month_year, last_month),
            "last_user_mean": self.get_month_mean(last_month_year, last_month, user=request.user),
            "user_monthly_hours": user_monthly_hours,
            "esfa_monthly_hours": self.get_esfa_monthly_means(year),
            "current_year": year,
            "current_month": month,
        }

        return Response(info, status=status.HTTP_200_OK)

    def get_hero(self, year: int, month: int) -> str:
        hero_name = "Anonymous Anonymousian"
        hero_row = (
            Sheet.objects
            .filter(year=year, month=month, user__is_active=True)
            .values("user__first_name", "user__last_name", "user_name")
            .order_by("-total")
            .first()
        )
        if hero_row:
            full_name = f"{hero_row.get('user__first_name', '')} {hero_row.get('user__last_name', '')}".strip()
            hero_name = full_name or hero_row.get("user_name") or hero_name
        return hero_name

    def get_month_mean(self, year: int, month: int, user=None) -> float:
        sheets = Sheet.objects.filter(year=year, month=month, user__is_active=True)
        if user is not None:
            sheets = sheets.filter(user=user)
        totals = sheets.aggregate(total_sum=Sum("total"), users_count=Count("id"))
        users_count = totals.get("users_count") or 0
        if not users_count:
            return 0
        return (totals.get("total_sum") or 0) / users_count

    def get_esfa_monthly_means(self, year: int) -> list:
        """Return monthly ESFA mean hours in one cheap aggregate query.

        The old page made 12 sequential calls to the public monthly report API. Each
        call transformed every monthly sheet with pandas, which made /hours_info
        very slow and could timeout. This keeps the chart data in the main payload.
        """
        rows = (
            Sheet.objects
            .filter(year=year, user__is_active=True)
            .values("month")
            .annotate(total_sum=Sum("total"), active_users=Count("id"))
            .order_by("month")
        )
        data = []
        for row in rows:
            active_users = row.get("active_users") or 0
            total_sum = row.get("total_sum") or 0
            if not active_users or total_sum <= 0:
                continue
            data.append({
                "month": row["month"],
                "meanHours": total_sum / active_users / 60,
            })
        return data

    @classmethod
    def hhmm2minutes(cls, value) -> int:
        try:
            if value is None:
                return 0
            if isinstance(value, (int, float)):
                return int(value)
            h, m = str(value).split(":")[:2]
            return int(h) * 60 + int(m)
        except Exception:
            return 0

    @classmethod
    def parse_project_prop(cls, value) -> float:
        try:
            if value is None or value == "":
                return 0
            if isinstance(value, str):
                return float(value.replace("%", "").strip()) / 100
            return float(value)
        except Exception:
            return 0

    @classmethod
    def row_hours(cls, row: dict, sheet: Sheet) -> int:
        auto_m = cls.hhmm2minutes(row.get("Auto Hours", "00:00"))
        remote_m = cls.hhmm2minutes(row.get("Remote", "00:00"))
        rest_m = cls.hhmm2minutes(row.get("Rest", "00:00"))
        computed = auto_m + remote_m - rest_m
        # Old rows may have only Hours. Also keep submitted manual Hours when no
        # attendance/remote/rest data exists yet.
        if computed == 0 and "Hours" in row:
            return cls.hhmm2minutes(row.get("Hours", "00:00"))
        return computed

    @classmethod
    def get_info(cls, queryset: QuerySet) -> pd.Series:
        totals = {}
        has_rows = False
        for sheet in queryset.iterator(chunk_size=200):
            for row in sheet.data or []:
                if not isinstance(row, dict):
                    continue
                has_rows = True
                hours = cls.row_hours(row, sheet)
                totals["Hours"] = totals.get("Hours", 0) + hours
                if not hours:
                    continue
                for key, value in row.items():
                    if key in cls.NON_PROJECT_COLUMNS:
                        continue
                    prop = cls.parse_project_prop(value)
                    if prop:
                        totals[key] = totals.get(key, 0) + (prop * hours)
        if not has_rows:
            return pd.Series(dtype="float64")
        return pd.Series(totals, dtype="float64")


class PublicMonthlyReportApiView(APIView):

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, year: str, month: str):

        sheets = Sheet.objects.filter(
            year=year, month=month, user__is_active=True
        ).only("id", "user", "year", "month", "data", "submitted")
        hours, activeUsers = PublicMonthlyReportApiView.get_sheet_sums(sheets)
        res = {
            "hours": hours,
            "activeUsers": activeUsers,
        }

        return Response(res, status=status.HTTP_200_OK)

    @classmethod
    def get_sheet_sums(cls, sheets: QuerySet) -> dict:
        projects = [p["name"] for p in Project.objects.values("name")]
        projects.append("Total")
        projects_sum = {p: 0 for p in projects}
        active_users = 0

        iterable = sheets.iterator(chunk_size=200) if hasattr(sheets, "iterator") else sheets
        for sheet in iterable:
            active_users += 1
            sheet_sum = cls.get_sum(sheet)
            for key, value in sheet_sum.items():
                projects_sum[key] = projects_sum.get(key, 0) + value

        return {
            key: cls.minute_formatter(value)
            for key, value in projects_sum.items()
        }, active_users

    @classmethod
    def get_sum(self, sheet: Sheet) -> dict:
        """Return sheet column sums without pandas; used by the public info page."""
        totals = {"Total": 0}
        for row in sheet.data or []:
            if not isinstance(row, dict):
                continue
            hours = InfoApiView.row_hours(row, sheet)
            totals["Total"] = totals.get("Total", 0) + hours
            if not hours:
                continue
            for key, value in row.items():
                if key in InfoApiView.NON_PROJECT_COLUMNS:
                    continue
                prop = InfoApiView.parse_project_prop(value)
                if prop:
                    totals[key] = totals.get(key, 0) + (prop * hours)
        return totals

    @classmethod
    def minute_formatter(cls, minutes: int) -> str:
        return f"{int(minutes // 60)}:{int(minutes % 60)}"


class MonthlyReportApiView(APIView):

    permission_classes = [customPermissions.IsProjectReportManager]

    REPORT_SHEET_FIELDS = [
        "id",
        "user",
        "year",
        "month",
        "data",
        "submitted",
        "is_verified",
        "is_supreme_verified",
    ]
    NEW_APPROVAL_FIELDS = [
        "manager_level_1_verified",
        "manager_level_2_verified",
        "supreme_verified",
    ]

    @classmethod
    def db_has_model_fields(cls, model, field_names):
        """Return False when code has new fields but the DB migration is not applied yet."""
        try:
            with connection.cursor() as cursor:
                table_columns = {
                    column.name
                    for column in connection.introspection.get_table_description(
                        cursor, model._meta.db_table
                    )
                }
        except Exception:
            return False

        for field_name in field_names:
            try:
                db_column = model._meta.get_field(field_name).column
            except Exception:
                return False
            if db_column not in table_columns:
                return False
        return True

    @classmethod
    def new_approval_columns_available(cls):
        return cls.db_has_model_fields(Sheet, cls.NEW_APPROVAL_FIELDS)

    @staticmethod
    def full_name_from_row(row):
        full_name = f"{row.get('first_name', '')} {row.get('last_name', '')}".strip()
        return full_name or "Unnamed User"

    @classmethod
    def get_active_user_report_rows(cls):
        # values() keeps this report working even if optional new User fields have
        # been added in code but their migration has not been applied yet.
        return list(
            User.objects
            .filter(is_active=True)
            .values("id", "first_name", "last_name", "wage")
        )

    @classmethod
    def get_report_sheet_queryset(cls, year, month, include_new_approval=None):
        if include_new_approval is None:
            include_new_approval = cls.new_approval_columns_available()

        fields = list(cls.REPORT_SHEET_FIELDS)
        if include_new_approval:
            fields.extend(cls.NEW_APPROVAL_FIELDS)

        # Do not select_related('user') here. If the new User manager fields are
        # not migrated yet, selecting the full User row can also crash reports.
        return (
            Sheet.objects
            .filter(year=year, month=month, user__is_active=True)
            .only(*fields)
        )

    def get(self, request, year: str, month: str):
        has_new_approval_columns = self.new_approval_columns_available()
        sheets_list = list(
            self.get_report_sheet_queryset(
                year, month, include_new_approval=has_new_approval_columns
            )
        )

        active_user_rows = self.get_active_user_report_rows()
        user_name_map = {
            row["id"]: self.full_name_from_row(row) for row in active_user_rows
        }
        user_wage_map = {row["id"]: row.get("wage", 0) or 0 for row in active_user_rows}
        sheet_user_ids = {sheet.user_id for sheet in sheets_list if sheet.user_id}
        sheetless_rows = [
            row for row in active_user_rows if row["id"] not in sheet_user_ids
        ]

        def sheet_user_name(sheet):
            return user_name_map.get(sheet.user_id, "Deleted User")

        submitted_names = [
            sheet_user_name(sheet) for sheet in sheets_list if sheet.submitted
        ]
        if has_new_approval_columns:
            verified_names = [
                sheet_user_name(sheet)
                for sheet in sheets_list
                if sheet.manager_level_1_verified and sheet.manager_level_2_verified
            ]
            supreme_verified_names = [
                sheet_user_name(sheet) for sheet in sheets_list if sheet.supreme_verified
            ]
        else:
            # Legacy fallback when the approval migration has not been applied.
            verified_names = [
                sheet_user_name(sheet) for sheet in sheets_list if sheet.is_verified
            ]
            supreme_verified_names = [
                sheet_user_name(sheet) for sheet in sheets_list if sheet.is_supreme_verified
            ]

        sheetless_names = [self.full_name_from_row(row) for row in sheetless_rows]

        hours, payments = self.get_sheet_sums(
            sheets_list,
            sheetless_rows,
            user_name_map=user_name_map,
            user_wage_map=user_wage_map,
        )

        res = {
            "hours": hours,
            # "payments": payments,
            "usersNum": len(active_user_rows),
            "sheetsNum": len(sheets_list),
            "submittedSheetsNum": len(submitted_names),
            "sheetlessUsers": sheetless_names,
            "submittedUsers": submitted_names,
            "verifiedUsers": verified_names,
            "suprimeVerifiedUsers": supreme_verified_names,
        }

        return Response(res, status=status.HTTP_200_OK)

    @classmethod
    def report_row_name(cls, row_or_user):
        if isinstance(row_or_user, dict):
            return cls.full_name_from_row(row_or_user)
        return row_or_user.get_full_name()

    @classmethod
    def normalize_sheet_data_for_report(cls, sheet: Sheet):
        """Normalize old row shapes in memory only; reports must not write sheets."""
        data = sheet.data or []
        if not data:
            return

        if "Remote" not in data[0]:
            for row in data:
                row.setdefault("Remote", "00:00")
                row.setdefault("Auto Hours", "00:00")
                row.setdefault("Rest", "00:00")

        for row in data:
            row.setdefault("Note Hours", "")
            auto_m = sheet.hhmm2minutes(row.get("Auto Hours", "00:00"))
            remote_m = sheet.hhmm2minutes(row.get("Remote", "00:00"))
            rest_m = sheet.hhmm2minutes(row.get("Rest", "00:00"))
            row["Total"] = sheet.minutes2hhmm(auto_m + remote_m - rest_m)

    @classmethod
    def get_sheet_sums(
        cls,
        sheets: QuerySet,
        sheetless_users: QuerySet,
        user_name_map=None,
        user_wage_map=None,
    ) -> dict:
        user_name_map = user_name_map or {}
        user_wage_map = user_wage_map or {}
        projects = [p["name"] for p in Project.objects.values("name")]
        projects.append("Total")
        # projects.append("Total 2")

        # a pandas Series which contains all projects.
        # a user's sheet sum should be added to this Series in order to contain all projects even the value is 0
        projects_empty = pd.Series({p: 0 for p in projects})
        # a pandas Series which will be a summation of all desiered sheet sums
        projects_sum = pd.Series({p: 0 for p in projects})

        hours = dict()
        payments = dict()
        for sheet in sheets:
            cls.normalize_sheet_data_for_report(sheet)
            sheet_sum = cls.get_sum(sheet)
            sheet_sum = projects_empty.add(sheet_sum, fill_value=0)
            projects_sum = projects_sum.add(sheet_sum, fill_value=0)
            if user_name_map:
                full_name = user_name_map.get(sheet.user_id, "Deleted User")
                wage = user_wage_map.get(sheet.user_id, 0)
            else:
                full_name = sheet.user.get_full_name() if sheet.user else "Deleted User"
                wage = sheet.user.wage if sheet.user else 0
            hours[full_name] = sheet_sum.apply(cls.minute_formatter).to_dict()
            payments[full_name] = sheet_sum.apply(
                lambda x: int(x / 60 * wage)
            ).to_dict()
        for user in sheetless_users:
            full_name = cls.report_row_name(user)
            hours[full_name] = projects_empty.to_dict()
            payments[full_name] = 0
        hours["Total"] = projects_sum.apply(cls.minute_formatter).to_dict()
        return hours, payments

    @classmethod
    def get_sum(self, sheet: Sheet) -> pd.Series:
        """returns sheet column sums"""
        df = sheet.transform()
        if df.empty:
            return pd.Series(dtype="float64")
        df.drop(
            columns=[
                "Day", "WeekDay", "Attendance", "Description",
                "Auto Hours", "Rest", "Remote", "Total", "Note Hours"
            ],
            errors="ignore",
            inplace=True
        )
        df.rename(columns={"Hours": "Total"}, inplace=True)
        # Note/comment columns are strings. Keep only numeric project/hour data so
        # pandas does not concatenate text and then crash when totals are added.
        df = df.apply(pd.to_numeric, errors="coerce").fillna(0)
        return df.sum(numeric_only=True)

    @classmethod
    def minute_formatter(cls, minutes: int) -> str:
        return f"{int(minutes // 60)}:{int(minutes % 60)}"


class PaymentApiView(APIView):
    permission_classes = [customPermissions.IsPaymentManager]

    def get(self, request, year: str, month: str):
        users = User.objects.filter(is_active=1)
        data = list()

        for user in users:
            sheet, created = Sheet.objects.get_or_create(
                user=user, year=year, month=month
            )
            if created:
                sheet.setup_sheet()

            payment_info = sheet.get_payment_info()
            payment_info.update(
                {
                    "userID": user.id,
                    "user": user.get_full_name(),
                    "bankName": user.bank_name,
                }
            )
            data.append(payment_info)

        return Response(data, status=status.HTTP_200_OK)

    def post(self, request, year: str, month: str):
        editted_row = request.data["row"]
        id = editted_row["userID"]
        wage = editted_row["wage"]
        base = editted_row["basePayment"]
        r1 = int(editted_row["reduction1"])
        r2 = int(editted_row["reduction2"])
        add1 = int(editted_row["addition1"])

        user = User.objects.get(pk=id)
        user.wage = wage
        user.base_payment = base
        user.reduction1 = r1
        user.reduction2 = r2
        user.addition1 = add1
        user.save()
        user_sheets = Sheet.objects.filter(user=id, year=year)
        for sheet in user_sheets:
            if sheet.month >= int(month):
                sheet.wage = wage
                sheet.base_payment = base
                sheet.reduction1 = r1
                sheet.reduction2 = r2
                sheet.addition1 = add1
                sheet.save()

        currentSheet = Sheet.objects.get(user=id, year=year, month=month)
        currentSheet.reduction3 = int(editted_row["reduction3"])
        currentSheet.food_reduction = int(editted_row["food_reduction"])
        currentSheet.addition2 = int(editted_row["addition2"])
        currentSheet.payment_status = int(editted_row["paymentStatus"])
        currentSheet.save()

        return Response(currentSheet.get_payment_info(), status=status.HTTP_200_OK)


class PublicPaymentApiView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, year: str, month: str):
        sheet = Sheet.objects.get(user=self.request.user, year=year, month=month)
        data = sheet.get_public_payment_info()
        return Response(data, status=status.HTTP_200_OK)


class OrderFoodApiView(APIView):

    def get(self, request, year: str, month: str):
        sheet, created = Sheet.objects.get_or_create(
            user=self.request.user, year=year, month=month
        )
        if created:
                sheet.setup_sheet()

        return Response(sheet.food_data, status=status.HTTP_200_OK)

    def post(self, request, year: str, month: str):
        data = request.data["data"]
        index = int(request.data["index"])
        sheet = Sheet.objects.get(user=self.request.user, year=year, month=month)
        self.updateSheetFoodData(data, index, sheet)
        self.update_sheet_food_reduction(sheet, year, month)
        return Response(sheet.food_data, status=status.HTTP_200_OK)

    def updateSheetFoodData(self, data, i, sheet):
        if i >= len(sheet.food_data):
            sheet.food_data.extend([[]] * (i + 1 - len(sheet.food_data)))
        sheet.food_data[i] = data
        sheet.save()

    def update_sheet_food_reduction(self, sheet, year, month):
        food_data, created = Food_data.objects.get_or_create(year=year, month=month)

        flat_list = [item for sublist in sheet.food_data for item in sublist]
        order_data = [
            item
            for item in flat_list
            if item["month"] == month and len(item["foods"]) > 0
        ]

        next_month = f"{(int(month)+1)%12}"
        next_year = year
        if next_month == "1":
            next_year = f"{int(year)+1}"

        next_month_order_data = [
            item
            for item in flat_list
            if item["month"] == next_month and len(item["foods"]) > 0
        ]
        if next_month_order_data != []:
            next_food_data, _ = Food_data.objects.get_or_create(
                year=next_year, month=next_month
            )
            next_sheet, created = Sheet.objects.get_or_create(
                user_id=sheet.user_id, year=next_year, month=next_month
            )
            if created:
                next_sheet.setup_sheet()

            next_sheet.food_reduction = self.calculateSheetFoodPrice(
                next_month_order_data, next_food_data
            )
            next_sheet.save()

        sheet.food_reduction = self.calculateSheetFoodPrice(order_data, food_data)
        sheet.save()

    def calculateSheetFoodPrice(self, order_data, food_data: Food_data):
        food_price_map = self.get_food_price_map(food_data.data)

        # Step 2: Calculate total price
        total_price = 0

        for order in order_data:
            order_day = int(order["day"])
            foods = order["foods"]
            today_price = 0
            today_price += self.get_delivery_price(
                food_data.statistics_and_cost_data, order_day
            )

            # Determine the applicable day for pricing
            applicable_day = 1
            for day in food_data.data:
                if day["day"] <= order_day:
                    applicable_day = day["day"]
                else:
                    break

            for food_id in foods:
                food_price_key = (applicable_day, food_id)
                if food_price_key in food_price_map:
                    today_price += food_price_map[food_price_key]
                else:
                    raise KeyError(
                        "food key error in OrderFoodApi.calculateSjeetFoodData"
                    )
                    continue
            total_price += today_price

        return total_price

    @classmethod
    def get_food_price_map(self, food_data):
        food_price_map = {}

        # Iterate over the food data and map prices based on the day ranges
        for food_entry in food_data:
            day = food_entry["day"]
            for food_item in food_entry["data"]:
                food_id = food_item["id"]
                price = food_item["price"]
                food_price_map[(day, food_id)] = price
        return food_price_map

    def get_delivery_price(self, cost_data, day):
        num_users_ordered = cost_data[day - 1]["num_users_Ordered"]
        whole_delivery_cost = cost_data[day - 1]["delivery_cost"]
        if num_users_ordered == 0:
            return whole_delivery_cost
        return whole_delivery_cost / num_users_ordered


class FoodDataApiView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, year: str, month: str):
        food_data = FoodManagementApiView.get_or_create_food_data(year, month)
        return Response(
            [food_data.data, food_data.order_mode], status=status.HTTP_200_OK
        )


class FoodManagementApiView(APIView):
    permission_classes = [customPermissions.IsFoodManager]

    def get(self, request, year: str, month: str):
        food_data = self.get_or_create_food_data(year, month)

        return Response(
            [food_data.data, food_data.order_mode, food_data.statistics_and_cost_data],
            status=status.HTTP_200_OK,
        )

    @classmethod
    def get_or_create_food_data(self, year, month):
        last_food_data = (
            Food_data.objects.exclude(data__exact=[])
            .order_by("-year", "-month")
            .first()
        )
        food_data, created = Food_data.objects.get_or_create(year=year, month=month)
        if created or food_data.data == []:
            if last_food_data is not None and len(last_food_data.data) > 0:
                newfooddata = last_food_data.data[-1]
                newfooddata["day"] = 1
                food_data.data = [newfooddata]
                food_data.save()
        if food_data.statistics_and_cost_data == []:
            if last_food_data.statistics_and_cost_data:
                default_delivery_price = last_food_data.statistics_and_cost_data[-1][
                    "delivery_cost"
                ]
            else:
                default_delivery_price = 200000
            empty_sheet = Sheet.empty_sheet_data(int(year), int(month))
            food_data.statistics_and_cost_data = [
                {
                    "day": item["Day"],
                    "num_users_Ordered": 0,
                    "delivery_cost": default_delivery_price,
                    "calculated_amount": 0,
                    "amount_paid": 0,
                }
                for item in empty_sheet
            ]
            food_data.save()
        return food_data

    def post(self, request, year: str, month: str):
        submittedFoodNames = request.data
        food_data = Food_data.objects.get(year=year, month=month)

        if not food_data.data:
            food_data.data = [
                {
                    "day": 1,
                    "data": [
                        {"id": int(key), "name": name, "price": 0}
                        for key, name in submittedFoodNames.items()
                    ],
                }
            ]
        else:
            submitted_ids = {int(key) for key in submittedFoodNames.keys()}

            for db_record in food_data.data:
                existing_items = {item["id"]: item for item in db_record["data"]}

                for key, name in submittedFoodNames.items():
                    food_id = int(key)
                    existing_item = existing_items.get(food_id)
                    if existing_item:
                        existing_item["name"] = name
                    else:
                        db_record["data"].append(
                            {"id": food_id, "name": name, "price": 0}
                        )

                # Remove items from the database that are not in the submitted data
                db_record["data"] = [
                    item for item in db_record["data"] if item["id"] in submitted_ids
                ]

        food_data.save()

        return Response(food_data.data, status=status.HTTP_200_OK)

    def put(self, request, year: str, month: str):
        requestTargetType = request.data["type"]
        food_data = Food_data.objects.get(year=year, month=month)
        if requestTargetType == "order_mode":
            mode = request.data["data"]
            return self.updateOrderMode(food_data, mode)
        if requestTargetType == "food_data":
            submittedFoodData = request.data["data"]
            return self.updateFoodsPrice(year, month, submittedFoodData, food_data)

    def updateOrderMode(self, food_data, mode):
        food_data.order_mode = mode
        food_data.save()
        return Response({}, status=status.HTTP_200_OK)

    def updateFoodsPrice(self, year, month, submittedFoodData, food_data):
        if not food_data.data:
            food_data.data = [
                {
                    "day": 1,
                    "data": [
                        {"id": int(key), "name": name, "price": 0}
                        for key, name in submittedFoodData.items()
                    ],
                }
            ]
        else:
            food_data.data = submittedFoodData
        food_data.save()
        self.update_all_foodReductions(year, month)
        return Response(food_data.data, status=status.HTTP_200_OK)

    @classmethod
    def update_all_foodReductions(self, year, month):
        OrderFoodApiObject = OrderFoodApiView()

        sheets = Sheet.objects.filter(year=year, month=month, user__is_active=True)

        for sheet in sheets:
            sheet.food_reduction = 0
            if sheet.food_data != []:
                OrderFoodApiObject.update_sheet_food_reduction(sheet, year, month)
            sheet.save()

        return


class FoodCostManagementApiView(APIView):
    permission_classes = [customPermissions.IsFoodManager]

    def get(self, request, year: str, month: str):
        food_data = FoodManagementApiView.get_or_create_food_data(year, month)
        self.update_statistics_and_cost_data(food_data, year, month)
        return Response(food_data.statistics_and_cost_data, status=status.HTTP_200_OK)

    def post(self, request, year: str, month: str):
        index, row_data = request.data.values()
        food_data = Food_data.objects.get(year=year, month=month)
        old_row_data = food_data.statistics_and_cost_data[index]
        if old_row_data["delivery_cost"] == row_data["delivery_cost"]:
            row_data["amount_paid"] = int(row_data["amount_paid"])
            food_data.statistics_and_cost_data[index] = row_data
            food_data.save()
        else:
            self.update_foods_delivery_cost(food_data, index, row_data["delivery_cost"])
            FoodManagementApiView.update_all_foodReductions(year, month)

        res = food_data.statistics_and_cost_data
        return Response(res, status=status.HTTP_200_OK)

    def update_foods_delivery_cost(self, food_data, idx, delivery_cost):
        for i in range(idx, len(food_data.statistics_and_cost_data)):
            food_data.statistics_and_cost_data[i]["delivery_cost"] = int(delivery_cost)
        food_data.save()

    def update_statistics_and_cost_data(self, food_data: Food_data, year, month):
        sheets = self.get_now_and_previous_month_sheets(year, month)

        for item in food_data.statistics_and_cost_data:
            item["calculated_amount"] = item["delivery_cost"]
            item["num_users_Ordered"] = 0
            # item['amount_paid'] = 0

        food_price_map = OrderFoodApiView.get_food_price_map(food_data.data)
        for sh in sheets:
            if sh.food_data:
                result = {
                    int(item["day"]): item["foods"]
                    for sublist in sh.food_data
                    for item in sublist
                    if item["month"] == month and len(item["foods"]) > 0
                }
                for day, food_ids in result.items():
                    food_data.statistics_and_cost_data[day - 1][
                        "num_users_Ordered"
                    ] += 1
                    try:
                        food_data.statistics_and_cost_data[day - 1][
                            "calculated_amount"
                        ] += self.get_order_price(
                            food_price_map, food_data, food_ids, day
                        )
                    except KeyError as e:
                        self.deselect_invalid_food(sh.food_data, e.args[0][1], month)
                        sh.save()
                        self.update_statistics_and_cost_data(food_data, year, month)
                        return
        self.handleFakeDeliveryPrices(food_data.statistics_and_cost_data)
        food_data.save()

    def deselect_invalid_food(self, food_monthly_data, food_id, target_month):
        for week_data in food_monthly_data:
            for day_data in week_data:
                if day_data["month"] == target_month and food_id in day_data["foods"]:
                    day_data["foods"].remove(food_id)

    def handleFakeDeliveryPrices(self, statistics_and_cost_data):
        for item in statistics_and_cost_data:
            if item["num_users_Ordered"] == 0:
                item["calculated_amount"] = 0

    @classmethod
    def get_now_and_previous_month_sheets(self, year, month):
        month = int(month)
        year = int(year)
        previous_month = month - 1 if month > 1 else 12
        previous_month_year = year if month > 1 else year - 1

        # Filter objects
        sheets = Sheet.objects.filter(
            Q(year=year, month=month)
            | Q(year=previous_month_year, month=previous_month)
        ).exclude(food_data=[])

        return sheets

    @classmethod
    def get_order_price(
        self, food_price_map, food_data: Food_data, food_ids, order_day
    ):
        applicable_day = 1
        for day in food_data.data:
            if day["day"] <= order_day:
                applicable_day = day["day"]
            else:
                break
        total_price = 0
        for food_id in food_ids:
            food_price_key = (applicable_day, food_id)
            # if food_price_key in food_price_map:
            total_price += food_price_map[food_price_key]
            # else:
            #     raise KeyError("food key error in get_order_price")
            #     continue
        return total_price


class DailyFoodsOrder(APIView):
    permission_classes = [customPermissions.IsFoodManager]

    def get(self, request, year: str, month: str, weekIndex: str, day: str):
        sheets = Sheet.objects.filter(year=year, month=month, user__is_active=True).exclude(food_data=[])
        food_data = Food_data.objects.get(year=year, month=month)
        if len(food_data.data) == 0:
            return Response([], status=status.HTTP_200_OK)
        food_data = food_data.data[0]["data"]

        d = [
            {"id": item["id"], "name": item["name"], "count": 0, "users": []}
            for item in food_data
        ]

        for sh in sheets:
            if sh.food_data is not [] and len(sh.food_data) > int(weekIndex):
                TargetWeekFoodData = sh.food_data[int(weekIndex)]
                selectedFoods = next(
                    (item for item in TargetWeekFoodData if item["day"] == day), None
                )
                if selectedFoods:
                    name = sh.user_name
                    self.update_counts(d, selectedFoods["foods"], name)

        return Response(d, status=status.HTTP_200_OK)

    def post(self, request, year: str, month: str, weekIndex: str, day: str):
        sheets = Sheet.objects.filter(year=year, month=month, user__is_active=True).exclude(food_data=[])
        food_data_obj = Food_data.objects.filter(year=year, month=month).first()
        if not food_data_obj or not food_data_obj.data:
            return Response([], status=status.HTTP_200_OK)

        food_data = food_data_obj.data[0]["data"]
        d = [
            [
                {"id": item["id"], "name": item["name"], "count": 0, "users": []}
                for item in food_data
            ]
            for _ in range(7)
        ]

        for sh in sheets:
            if sh.food_data is not [] and len(sh.food_data) > int(weekIndex):
                TargetWeekFoodData = sh.food_data[int(weekIndex)]
                for index, item in enumerate(TargetWeekFoodData):
                    self.update_counts(d[index], item["foods"], sh.user_name)

        unique_food_names = list(
            OrderedDict.fromkeys(item["name"] for item in food_data)
        )
        weekdays = [
            "شنبه",
            "یکشنبه",
            "دوشنبه",
            "سه شنبه",
            "چهارشنبه",
            "پنچ شنبه",
            "جمعه",
        ]
        food_data_to_order = {"روز/غذا": weekdays}
        food_data_to_order.update({name: [0] * 7 for name in unique_food_names})
        food_data_to_order.update({"مجموع": [0] * 7})
        for day_idx, day_data in enumerate(d):
            food_count_map = {item["name"]: item["count"] for item in day_data}
            for name in unique_food_names:
                food_data_to_order[name][day_idx] = food_count_map.get(name, 0)
                food_data_to_order["مجموع"][day_idx] += food_count_map.get(name, 0)
                continue

        food_user_data = {"*": {i: day for i, day in enumerate(weekdays)}}
        for list_index, sublist in enumerate(d):
            # Iterate through each item in the sublist
            for item in sublist:
                food_name = item["name"]
                for user in item["users"]:
                    if user not in food_user_data:
                        food_user_data[user] = {}
                    if list_index in food_user_data[user]:
                        food_user_data[user][list_index] += f", {food_name}"
                    else:
                        food_user_data[user][list_index] = food_name

        df = pd.DataFrame(food_data_to_order)
        df2 = pd.DataFrame(food_user_data)
        df_transposed = df2.transpose()

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False)
            df_transposed.to_excel(writer, sheet_name="Sheet2", index=True)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f"attachment; filename=food_order_{year}_{month}_week{weekIndex}.xlsx"
        )
        return response

    def update_counts(self, items, ids, user):
        for item in items:
            if item["id"] in ids:
                item["count"] += 1
                item["users"].append(user)


class DailyReportUser(APIView):
    def get(self, request, year: str, month: str, day: str):
        user = request.user

        reportSetting, _ = DailyReportSetting.objects.get_or_create()

        report = Report.objects.filter(
            user=user, year=year, month=month, day=day
        ).first()

        res = {
            "content": report.content if report else "",
            "main_comment": (
                report.main_comment
                if report and not report.manager_comment_hide_for_user
                else ""
            ),
            "sub_comment": (
                report.sub_comment
                if report and not report.supervisor_comment_hide_for_user
                else ""
            ),
            "no_limit_submit_btn": reportSetting.no_limit_submission,
            "start_report_hour": reportSetting.start_report_hour,
            "end_report_hour": reportSetting.end_report_hour,
        }

        return Response(res, status=status.HTTP_200_OK)

    def post(self, request, year: str, month: str, day: str):
        user = request.user
        data = request.data
        content = data.get("content")

        _, created = Report.objects.update_or_create(
            user=user,
            year=year,
            month=month,
            day=day,
            defaults={"content": content},
        )

        return Response(
            {
                "message": (
                    "Report created successfully"
                    if created
                    else "Report updated successfully"
                )
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )


class DailyReportManagement(APIView):
    permission_classes = [customPermissions.IsDailyReportManager]

    def get(self, request, year: str, month: str):
        reports = Report.objects.filter(year=year, month=month).order_by("user", "day")

        # Group reports by user
        grouped_reports = {}
        for username, items in groupby(
            reports, key=lambda report: report.user.username
        ):
            grouped_reports[username] = [
                {
                    "day": report.day,
                    "content": report.content,
                    "manager_comment": report.main_comment,
                    "manager_comment_hide_for_user": report.manager_comment_hide_for_user,
                    "manager_comment_hide_for_supervisor": report.manager_comment_hide_for_supervisor,
                    "supervisor_comment": report.sub_comment,
                    "supervisor_comment_hide_for_user": report.supervisor_comment_hide_for_user,
                }
                for report in items
            ]

        return Response(
            {
                "user": {
                    "id": request.user.id,
                    "is_MainReportManager": request.user.is_MainReportManager,
                },
                "data": grouped_reports,
            },
            status=status.HTTP_200_OK,
        )

    def post(self, request, year: str, month: str):
        data = request.data
        user = data["userName"]
        day = str(data["day"])
        is_manager_submitted = data["is_manager_submitted"]

        # update corresponding report object
        report, created = Report.objects.get_or_create(
            user__username=user, year=year, month=month, day=day
        )
        if created:
            userObj = User.objects.get(username=user)
            report.user = userObj

        if is_manager_submitted:
            report.main_comment = data["manager_comment"]
            report.manager_comment_hide_for_user = data["manager_comment_hide_for_user"]
            report.manager_comment_hide_for_supervisor = data[
                "manager_comment_hide_for_supervisor"
            ]
        else:
            report.sub_comment = data["supervisor_comment"]
            report.supervisor_comment_hide_for_user = data[
                "supervisor_comment_hide_for_user"
            ]

        report.save()
        return Response(
            {"message": "Report updated successfully"}, status=status.HTTP_200_OK
        )

class ExportDailyReportManagement(APIView):
    permission_classes = [customPermissions.IsDailyReportManager]

    def post(self, request, year: str, month: str):
        year_int = int(year)
        month_int = int(month)
        # Get all days in the month (even without reports)

        is_leap = jdt.date(year_int, month_int, 1).isleap()
        month_days = current_mont_days(month_int, is_leap)
        
        reports = Report.objects.filter(year=year, month=month).order_by("user", "day")
        reports_by_user = {}
        
        # Create report dictionary with all days
        for report in reports:
            username = report.user.username if report.user else "Unknown User"
            if username not in reports_by_user:
                # Initialize with empty entries for all days
                reports_by_user[username] = []
                for day in range(1, month_days + 1):
                    try:
                        jalali_date = jdt.date(year_int, month_int, day)
                        reports_by_user[username].append({
                            'date': jalali_date,
                            'date_display': jalali_date.strftime("%d %B %Y"),
                            'weekday': get_persian_weekday(jalali_date),
                            'is_thursday': jalali_date.weekday() == 5,
                            'is_friday': jalali_date.weekday() == 6,
                            'content': '',
                            "Manager's Comment": '',
                            "Supervisor's Comment": '',
                            'day': day
                        })
                    except:
                        continue
            
            # Update with actual report data
            for entry in reports_by_user[username]:
                if entry['day'] == report.day:
                    entry.update({
                        'content': report.content,
                        "Manager's Comment": report.main_comment,
                        "Supervisor's Comment": report.sub_comment,
                    })
                    break
        
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            used_sheet_names = set()
            
            for username, user_reports in reports_by_user.items():
                # Convert to DataFrame for easier handling
                df_data = []
                for report in user_reports:
                    df_data.append({
                        'تاریخ': report['date_display'],
                        'روز هفته': report['weekday'],
                        'گزارش': report['content'],
                        'نظر مدیر': report["Manager's Comment"],
                        'نظر سرپرست': report["Supervisor's Comment"],
                        'is_thursday': report['is_thursday'],  # Temp column for formatting
                        'is_friday': report['is_friday']  # Temp column for formatting
                    })
                
                df = pd.DataFrame(df_data)
                
                sheet_name = username[:31]
                counter = 1
                while sheet_name.lower() in used_sheet_names:
                    sheet_name = f"{username[:28]}_{counter}"
                    counter += 1
                
                used_sheet_names.add(sheet_name.lower())
                
                # Write DataFrame to Excel
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                
                # Apply styling
                self._apply_excel_styling(worksheet, df)
        
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=reports_{year}_{month}.xlsx'
        
        return response

    def _apply_excel_styling(self, worksheet, df):
        # Define styles
        persian_font = Font(name='Calibri', size=11)
        header_font = Font(name='Calibri', size=14, bold=True)
        
        # Different alignments for different columns
        header_alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        rtl_alignment = Alignment(vertical='center', wrap_text=True, horizontal='right', readingOrder=2)
        center_alignment = Alignment(vertical='center', wrap_text=True, horizontal='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        friday_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        thursday_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Apply basic styling to all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = persian_font
                cell.border = thin_border
        
        # Apply header styling with center alignment
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = header_alignment
        
        # Apply specific alignments to data rows
        max_row = len(df) + 1
        for row_idx in range(2, max_row + 1):
            # Columns A & B (تاریخ and روز هفته) - LTR alignment
            worksheet.cell(row=row_idx, column=1).alignment = center_alignment  # تاریخ
            worksheet.cell(row=row_idx, column=2).alignment = center_alignment  # روز هفته
            
            # Columns C, D, E (محتوا, نظر مدیر, نظر سرپرست) - RTL alignment and reading order
            for col_idx in [3, 4, 5]:
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = rtl_alignment
        
        # Apply Friday coloring only to first two cells (date and weekday)
        for row_idx in range(2, max_row + 1):
            is_thursday = worksheet.cell(row=row_idx, column=6).value  # Hidden is_thursday column
            is_friday = worksheet.cell(row=row_idx, column=7).value  # Hidden is_friday column
            if is_friday:
                worksheet.cell(row=row_idx, column=1).fill = friday_fill  # تاریخ
                worksheet.cell(row=row_idx, column=2).fill = friday_fill  # روز هفته
            if is_thursday:
                worksheet.cell(row=row_idx, column=1).fill = thursday_fill  # تاریخ
                worksheet.cell(row=row_idx, column=2).fill = thursday_fill  # روز هفته
        
        # Delete columns F and G
        worksheet.delete_cols(6, 2)  

        
        # Adjust column widths with wrap text
        column_widths = {
            'A': 15,  # تاریخ
            'B': 15,  # روز هفته
            'C': 60,  # محتوا
            'D': 40,  # نظر مدیر
            'E': 30   # نظر سرپرست
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Create Excel table
        table_ref = f"A1:E{max_row}"
        table = Table(displayName=f"Table_{worksheet.title}", ref=table_ref)
        
        # Use blue table style (Medium 6 is blue style)
        style = TableStyleInfo(
            name="TableStyleMedium6",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        worksheet.add_table(table)
        
        # Freeze header row
        worksheet.freeze_panes = "A2"


class DailyReportSettingManager(APIView):
    permission_classes = [customPermissions.IsDailyReportManager]

    def get(self, request):
        settign, _ = DailyReportSetting.objects.get_or_create()
        return Response(
            {"no_limit_submission": settign.no_limit_submission},
            status=status.HTTP_200_OK,
        )

    def post(self, request):
        data = request.data
        limitation = data["no_limit_submission"]
        settign, _ = DailyReportSetting.objects.get_or_create()

        settign.no_limit_submission = limitation
        settign.save()
        return Response(
            {"message": "Report setting updated successfully"},
            status=status.HTTP_200_OK,
        )


def _parse_verifier_group_tags(user):
    raw_tags = user.verifier_group_tags or ""
    try:
        return [int(tag.strip()) for tag in raw_tags.split(",") if tag.strip()]
    except ValueError:
        return []


def _has_legacy_group_access(verifier, staff):
    if verifier.is_SupremeHourVerifier:
        return True
    if not verifier.is_HourVerifier:
        return False
    return staff.staff_group_tag in _parse_verifier_group_tags(verifier)


def _is_manager_level_1(sheet, verifier):
    if not sheet.user:
        return False
    if sheet.user.manager_level_1_id == verifier.id:
        return True
    # Backward compatibility: old group/tag verifiers act as manager level 1
    # only where no direct manager_level_1 has been assigned yet.
    return sheet.user.manager_level_1_id is None and _has_legacy_group_access(verifier, sheet.user)


def _is_manager_level_2(sheet, verifier):
    return bool(sheet.user and sheet.user.manager_level_2_id == verifier.id)


def _can_view_sheet(sheet, verifier):
    return (
        verifier.is_SupremeHourVerifier
        or _is_manager_level_1(sheet, verifier)
        or _is_manager_level_2(sheet, verifier)
        or _has_legacy_group_access(verifier, sheet.user)
    )


def _can_verify_manager_level_1(sheet, verifier):
    return sheet.submitted and not sheet.manager_level_1_verified and _is_manager_level_1(sheet, verifier)


def _can_reject_manager_level_1(sheet, verifier):
    return (
        sheet.submitted
        and _is_manager_level_1(sheet, verifier)
        and not sheet.manager_level_2_verified
        and not sheet.supreme_verified
    )


def _can_verify_manager_level_2(sheet, verifier):
    # Sequential approval assumption: manager level 2 acts after manager level 1.
    return (
        sheet.submitted
        and sheet.manager_level_1_verified
        and not sheet.manager_level_2_verified
        and _is_manager_level_2(sheet, verifier)
    )


def _can_reject_manager_level_2(sheet, verifier):
    return sheet.submitted and _is_manager_level_2(sheet, verifier) and not sheet.supreme_verified


def _can_verify_supreme(sheet, verifier):
    return sheet.submitted and verifier.is_SupremeHourVerifier and not sheet.supreme_verified


def _can_reject_supreme(sheet, verifier):
    return sheet.submitted and verifier.is_SupremeHourVerifier


def _sheet_summary(sheet, role):
    user = sheet.user
    auto_hours = 0
    for row in sheet.data:
        auto_hours += sheet.hhmm2minutes(row.get("Auto Hours", "00:00"))
    is_warning = auto_hours > 0 and sheet.total > 1.1 * auto_hours
    return {
        "userId": user.id,
        "userName": user.get_full_name(),
        "username": user.username,
        "staffGroup": user.staff_group_tag,
        "deviceCode": user.auto_hour_ID,
        "year": sheet.year,
        "month": sheet.month,
        "role": role,
        "isSubmitted": sheet.submitted,
        "managerLevel1Verified": sheet.manager_level_1_verified,
        "managerLevel2Verified": sheet.manager_level_2_verified,
        "supremeVerified": sheet.supreme_verified,
        "isVerified": sheet.manager_level_1_verified and sheet.manager_level_2_verified,
        "isSupremeVerified": sheet.supreme_verified,
        "isFullyApproved": sheet.is_fully_approved,
        "autoHours": auto_hours,
        "isWarning": is_warning,
        "totalHours": sheet.total,
        "managerLevel1Name": user.manager_level_1.get_full_name() if user.manager_level_1 else "",
        "managerLevel2Name": user.manager_level_2.get_full_name() if user.manager_level_2 else "",
        "lastRejectedAt": sheet.last_rejected_at.isoformat() if sheet.last_rejected_at else None,
        "rejectionReason": sheet.rejection_reason,
    }


def _apply_rejection(sheet, verifier, reason):
    sheet.reset_approval_state()
    sheet.last_rejected_by = verifier
    sheet.last_rejected_at = timezone.now()
    sheet.rejection_reason = reason or ""
    sheet.save()


class HourVerifierAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _has_manager_hour_access(self, verifier):
        if verifier.is_HourVerifier:
            return True
        return User.objects.filter(
            Q(manager_level_1=verifier) | Q(manager_level_2=verifier),
            is_active=True,
        ).exists()

    def _has_mode_access(self, verifier, mode):
        if mode == "supreme":
            return verifier.is_SupremeHourVerifier
        return self._has_manager_hour_access(verifier)

    def _can_view_sheet_for_mode(self, sheet, verifier, mode):
        if mode == "supreme":
            return verifier.is_SupremeHourVerifier
        return _is_manager_level_1(sheet, verifier) or _is_manager_level_2(sheet, verifier) or _has_legacy_group_access(verifier, sheet.user)

    def _base_sheets(self, year, month):
        return Sheet.objects.filter(
            year=year,
            month=month,
            user__is_active=True,
        ).select_related(
            "user",
            "user__manager_level_1",
            "user__manager_level_2",
            "manager_level_1_verified_by",
            "manager_level_2_verified_by",
            "supreme_verified_by",
            "last_rejected_by",
        )

    def _manager_sections(self, verifier, year, month):
        sheets = self._base_sheets(year, month)

        direct_filter = Q(user__manager_level_1=verifier) | Q(user__manager_level_2=verifier)
        legacy_tags = _parse_verifier_group_tags(verifier)
        if verifier.is_HourVerifier and legacy_tags:
            direct_filter |= Q(user__manager_level_1__isnull=True, user__staff_group_tag__in=legacy_tags)

        sheets = sheets.filter(direct_filter).order_by("user__last_name_p", "user__first_name_p", "user__username")
        sections = {"currentQueue": [], "other": [], "unsubmitted": [], "approved": []}

        for sheet in sheets:
            roles = []
            if _is_manager_level_1(sheet, verifier):
                roles.append("manager_level_1")
            if _is_manager_level_2(sheet, verifier):
                roles.append("manager_level_2")

            for role in roles:
                item = _sheet_summary(sheet, role)

                if not sheet.submitted:
                    # Rejected sheets must stay visible to managers, but clean
                    # never-submitted sheets belong in their own bucket.
                    if sheet.rejection_reason or sheet.last_rejected_at:
                        sections["other"].append(item)
                    else:
                        sections["unsubmitted"].append(item)
                    continue

                if role == "manager_level_1":
                    if not sheet.manager_level_1_verified:
                        sections["currentQueue"].append(item)
                    elif sheet.manager_level_1_verified:
                        sections["approved"].append(item)
                    else:
                        sections["other"].append(item)
                elif role == "manager_level_2":
                    # Sequential approval assumption.
                    if sheet.manager_level_1_verified and not sheet.manager_level_2_verified:
                        sections["currentQueue"].append(item)
                    elif sheet.manager_level_2_verified:
                        sections["approved"].append(item)
                    else:
                        sections["other"].append(item)

        return {
            "mode": "manager",
            "sections": sections,
            "labels": {
                "currentQueue": "Needs your approval",
                "other": "Other / not ready",
                "unsubmitted": "Unsubmitted",
                "approved": "Approved by you",
            },
            "sequentialManagerLevel2": True,
        }

    def _supreme_sections(self, year, month):
        sections = {"currentQueue": [], "other": [], "unsubmitted": [], "approved": []}
        sheets = self._base_sheets(year, month).order_by(
            "user__last_name_p", "user__first_name_p", "user__username"
        )
        for sheet in sheets:
            item = _sheet_summary(sheet, "supreme")

            if not sheet.submitted:
                # Supreme rejection sets submitted=False, so include rejected
                # sheets here instead of letting them disappear from the panel.
                if sheet.rejection_reason or sheet.last_rejected_at:
                    sections["other"].append(item)
                else:
                    sections["unsubmitted"].append(item)
                continue

            if sheet.is_fully_approved:
                sections["approved"].append(item)
            elif not sheet.supreme_verified:
                sections["currentQueue"].append(item)
            else:
                sections["other"].append(item)

        return {
            "mode": "supreme",
            "sections": sections,
            "labels": {
                "currentQueue": "Submitted / waiting for supreme action",
                "other": "Other / rejected or still incomplete",
                "unsubmitted": "Unsubmitted",
                "approved": "Fully approved / ready for payment",
            },
            "sequentialManagerLevel2": True,
        }

    def get(self, request, year: str, month: str):
        verifier = request.user
        mode = request.query_params.get("mode", "manager")
        if mode not in ("manager", "supreme"):
            return Response({"error": "Invalid verification mode."}, status=status.HTTP_400_BAD_REQUEST)
        if not self._has_mode_access(verifier, mode):
            return Response({"error": "You do not have access to this verification panel."}, status=status.HTTP_403_FORBIDDEN)

        user_id = request.query_params.get("userId")
        if user_id:
            try:
                sheet = self._base_sheets(year, month).get(user_id=user_id)
            except Sheet.DoesNotExist:
                return Response({"error": "Sheet not found"}, status=status.HTTP_404_NOT_FOUND)

            if not self._can_view_sheet_for_mode(sheet, verifier, mode):
                return Response({"error": "You cannot view this sheet in this panel."}, status=status.HTTP_403_FORBIDDEN)

            auto_hours = 0
            total_hours = 0
            for row in sheet.data:
                auto_hours += sheet.hhmm2minutes(row.get("Auto Hours", "00:00"))
                total_hours += sheet.hhmm2minutes(row.get("Total", "00:00"))
            is_warning = auto_hours > 0 and total_hours > 1.1 * auto_hours

            can_edit_manager_1_comment = (mode == "supreme" and verifier.is_SupremeHourVerifier) or _is_manager_level_1(sheet, verifier)
            can_edit_manager_2_comment = (mode == "supreme" and verifier.is_SupremeHourVerifier) or _is_manager_level_2(sheet, verifier)

            data = _sheet_summary(sheet, request.query_params.get("role", ""))
            data.update({
                "sheetData": sheet.data,
                "autoHours": auto_hours,
                "totalHours": total_hours,
                "isWarning": is_warning,
                "managerLevel1Comment": sheet.manager_level_1_comment,
                "managerLevel2Comment": sheet.manager_level_2_comment,
                "rejectionReason": sheet.rejection_reason,
                "permissions": {
                    "canVerifyManagerLevel1": mode == "manager" and _can_verify_manager_level_1(sheet, verifier),
                    "canRejectManagerLevel1": mode == "manager" and _can_reject_manager_level_1(sheet, verifier),
                    "canVerifyManagerLevel2": mode == "manager" and _can_verify_manager_level_2(sheet, verifier),
                    "canRejectManagerLevel2": mode == "manager" and _can_reject_manager_level_2(sheet, verifier),
                    "canVerifySupreme": mode == "supreme" and _can_verify_supreme(sheet, verifier),
                    "canRejectSupreme": mode == "supreme" and _can_reject_supreme(sheet, verifier),
                    "canEditManagerLevel1Comment": can_edit_manager_1_comment,
                    "canEditManagerLevel2Comment": can_edit_manager_2_comment,
                },
            })
            return Response(data, status=status.HTTP_200_OK)

        if mode == "supreme":
            return Response(self._supreme_sections(year, month), status=status.HTTP_200_OK)

        return Response(self._manager_sections(verifier, year, month), status=status.HTTP_200_OK)

    def post(self, request, year: str, month: str):
        verifier = request.user
        mode = request.query_params.get("mode", request.data.get("mode", "manager"))
        user_id = request.data.get("userId")
        action = request.data.get("action")
        reason = request.data.get("reason", "")

        if mode not in ("manager", "supreme"):
            return Response({"error": "Invalid verification mode."}, status=status.HTTP_400_BAD_REQUEST)
        if not self._has_mode_access(verifier, mode):
            return Response({"error": "You do not have access to this verification panel."}, status=status.HTTP_403_FORBIDDEN)

        if not user_id:
            return Response({"error": "userId is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            sheet = self._base_sheets(year, month).get(user_id=user_id)
        except Sheet.DoesNotExist:
            return Response({"error": "Sheet not found"}, status=status.HTTP_404_NOT_FOUND)

        if not self._can_view_sheet_for_mode(sheet, verifier, mode):
            return Response({"error": "You cannot manage this sheet in this panel."}, status=status.HTTP_403_FORBIDDEN)

        # Allow comment edits to be bundled with verify/reject actions or saved alone.
        if "managerLevel1Comment" in request.data:
            if not ((mode == "supreme" and verifier.is_SupremeHourVerifier) or _is_manager_level_1(sheet, verifier)):
                return Response({"error": "You cannot edit manager level 1 comment."}, status=status.HTTP_403_FORBIDDEN)
            sheet.manager_level_1_comment = request.data.get("managerLevel1Comment", "")

        if "managerLevel2Comment" in request.data:
            if not ((mode == "supreme" and verifier.is_SupremeHourVerifier) or _is_manager_level_2(sheet, verifier)):
                return Response({"error": "You cannot edit manager level 2 comment."}, status=status.HTTP_403_FORBIDDEN)
            sheet.manager_level_2_comment = request.data.get("managerLevel2Comment", "")

        if action == "save_comments":
            sheet.save(update_fields=["manager_level_1_comment", "manager_level_2_comment"])
            return Response({"success": True}, status=status.HTTP_200_OK)

        if not sheet.submitted:
            return Response({"error": "Unsubmitted sheets cannot be verified or rejected."}, status=status.HTTP_400_BAD_REQUEST)

        now = timezone.now()

        if action and action.endswith("supreme") and mode != "supreme":
            return Response({"error": "Supreme actions must be performed from the supreme verifier panel."}, status=status.HTTP_400_BAD_REQUEST)
        if action and ("manager_level_1" in action or "manager_level_2" in action) and mode != "manager":
            return Response({"error": "Manager actions must be performed from the manager verifier panel."}, status=status.HTTP_400_BAD_REQUEST)

        if action == "verify_manager_level_1":
            if not _can_verify_manager_level_1(sheet, verifier):
                return Response({"error": "You cannot verify this sheet as manager level 1."}, status=status.HTTP_403_FORBIDDEN)
            sheet.manager_level_1_verified = True
            sheet.manager_level_1_verified_at = now
            sheet.manager_level_1_verified_by = verifier
            sheet.sync_legacy_verification_fields()
            sheet.save()
            return Response({"success": True}, status=status.HTTP_200_OK)

        if action == "reject_manager_level_1":
            if not _can_reject_manager_level_1(sheet, verifier):
                return Response({"error": "You cannot reject this sheet as manager level 1."}, status=status.HTTP_403_FORBIDDEN)
            _apply_rejection(sheet, verifier, reason)
            return Response({"success": True}, status=status.HTTP_200_OK)

        if action == "verify_manager_level_2":
            if not _can_verify_manager_level_2(sheet, verifier):
                return Response({"error": "You cannot verify this sheet as manager level 2."}, status=status.HTTP_403_FORBIDDEN)
            sheet.manager_level_2_verified = True
            sheet.manager_level_2_verified_at = now
            sheet.manager_level_2_verified_by = verifier
            sheet.sync_legacy_verification_fields()
            sheet.save()
            return Response({"success": True}, status=status.HTTP_200_OK)

        if action == "reject_manager_level_2":
            if not _can_reject_manager_level_2(sheet, verifier):
                return Response({"error": "You cannot reject this sheet as manager level 2."}, status=status.HTTP_403_FORBIDDEN)
            _apply_rejection(sheet, verifier, reason)
            return Response({"success": True}, status=status.HTTP_200_OK)

        if action == "verify_supreme":
            if not _can_verify_supreme(sheet, verifier):
                return Response({"error": "You cannot supreme-verify this sheet."}, status=status.HTTP_403_FORBIDDEN)
            sheet.supreme_verified = True
            sheet.supreme_verified_at = now
            sheet.supreme_verified_by = verifier
            sheet.sync_legacy_verification_fields()
            sheet.save()
            return Response({"success": True}, status=status.HTTP_200_OK)

        if action == "reject_supreme":
            if not _can_reject_supreme(sheet, verifier):
                return Response({"error": "You cannot reject this sheet as supreme verifier."}, status=status.HTTP_403_FORBIDDEN)
            _apply_rejection(sheet, verifier, reason)
            return Response({"success": True}, status=status.HTTP_200_OK)

        return Response({"error": "Unknown action"}, status=status.HTTP_400_BAD_REQUEST)
