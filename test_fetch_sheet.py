import pandas as pd
import numpy as np
import requests
import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def download_excel_file(url, local_file="debug_data.xlsx"):
    """Download Excel file from URL or load from local cache."""
    if os.path.exists(local_file):
        print(f"Loading from local cache: {local_file}")
        return local_file
    
    print(f"Fetching from web: {url}")
    try:
        response = requests.get(url)
        response.raise_for_status()
        
        with open(local_file, "wb") as f:
            f.write(response.content)
            
        return local_file
    except Exception as e:
        print(f"Failed to fetch data: {e}")
        raise


def parse_table_range(table_ref):
    """Parse Excel table range and return dimensions."""
    ref_parts = table_ref.split(':')
    if len(ref_parts) != 2:
        return None
    
    start_cell, end_cell = ref_parts
    
    # Extract column letters and row numbers
    start_col = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    end_col = ''.join(filter(str.isalpha, end_cell))
    end_row = int(''.join(filter(str.isdigit, end_cell)))
    
    # Convert to indices
    start_col_idx = column_index_from_string(start_col)
    end_col_idx = column_index_from_string(end_col)
    
    rows = end_row - start_row + 1
    cols = end_col_idx - start_col_idx + 1
    
    return {
        'range': table_ref,
        'start': start_cell,
        'end': end_cell,
        'start_row': start_row,
        'end_row': end_row,
        'start_col': start_col,
        'end_col': end_col,
        'start_col_idx': start_col_idx,
        'end_col_idx': end_col_idx,
        'rows': rows,
        'cols': cols,
        'data_rows': rows - 1
    }


def get_sheet_tables(file_path):
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet_tables = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            tables_info = []
            
            if hasattr(ws, 'tables'):
                for table_name, table_ref in ws.tables.items():
                    table_data = parse_table_range(table_ref)
                    if table_data:
                        table_data['name'] = table_name
                        tables_info.append(table_data)
            
            sheet_tables[sheet_name] = tables_info
        
        return sheet_tables
    
    except Exception as e:
        print(f"Error loading workbook: {e}")
        raise


def print_table_summary(sheet_tables):
    """Print formatted summary of tables in sheets."""
    print("Successfully fetched Excel file.")
    print("-" * 40)
    
    for sheet_name, tables in sheet_tables.items():
        print(f"### Sheet: '{sheet_name}'")
        print(f"   Found {len(tables)} Excel Table object(s):")
        
        for table in tables:
            print(f"\n   Table Name: '{table['name']}'")
            print(f"   Range: {table['range']}")
            print(f"   From {table['start']} to {table['end']}")
            print(f"   Dimensions: {table['rows']} rows (including header) x {table['cols']} columns")
            print(f"   Data rows: {table['data_rows']} rows (excluding header)")
        
        if not tables:
            print(f"   No formal Excel Table objects found")
        
        print()


def main():
    # Configure URL
    url = "https://docs.google.com/spreadsheets/d/1eFN39ZVBw7vNlX7N-XnhY7WtzvMfc6aCw_mu2fTufWs/edit?usp=sharing"
    url = url.replace("edit?usp=sharing", "export?format=xlsx")
    
    print(f"Attempting to fetch: {url}")
    
    # Download or load file
    local_file = download_excel_file(url)
    
    # Get table information
    sheet_tables = get_sheet_tables(local_file)
    
    # Print summary
    print_table_summary(sheet_tables)
    
    return sheet_tables


if __name__ == "__main__":
    result = main()